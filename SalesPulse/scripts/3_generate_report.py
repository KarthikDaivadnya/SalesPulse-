"""
SalesPulse - Step 3: Automated Report Generator
---------------------------------------------------
Generates a formatted Excel summary report (weekly or monthly) directly from
the cleaned dataset -- this is the script that would otherwise require a
person to manually pull numbers and rebuild a report every week/month.

Usage:
    python scripts/3_generate_report.py --period monthly
    python scripts/3_generate_report.py --period weekly

Each run reads the latest data and regenerates the report from scratch, so it
can be safely scheduled (e.g. via cron / Task Scheduler) without manual
intervention.
"""
import argparse
import pandas as pd
from datetime import datetime
from openpyxl import Workbook  # type: ignore[import]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import]
from openpyxl.chart import LineChart, Reference  # type: ignore[import]

NAVY = "1F3864"
ACCENT_BLUE = "2E5FAC"
LIGHT_BLUE = "D9E2F3"
GREY = "F2F2F2"

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=16)
SUBTITLE_FONT = Font(name="Calibri", italic=True, color="555555", size=10)
LABEL_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=11)
KPI_VALUE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=20)
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="555555")

HEADER_FILL = PatternFill("solid", start_color=ACCENT_BLUE)
KPI_FILL = PatternFill("solid", start_color=LIGHT_BLUE)
ALT_ROW_FILL = PatternFill("solid", start_color=GREY)

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def write_table(ws, df, start_row, start_col=1, currency_cols=None, pct_cols=None):
    currency_cols = currency_cols or []
    pct_cols = pct_cols or []
    for j, col in enumerate(df.columns):
        ws.cell(row=start_row, column=start_col + j, value=col.replace("_", " ").title())
    style_header_row(ws, start_row, len(df.columns), start_col)

    for i, (_, row) in enumerate(df.iterrows()):
        r = start_row + 1 + i
        fill = ALT_ROW_FILL if i % 2 == 1 else None
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=r, column=start_col + j, value=row[col])
            cell.font = BODY_FONT
            cell.border = BORDER
            if fill:
                cell.fill = fill
            if col in currency_cols:
                cell.number_format = '"\u20b9"#,##0'
            elif col in pct_cols:
                cell.number_format = "0.0%"
    return start_row + 1 + len(df)


def add_kpi_card(ws, col, row, label, value, fmt="number"):
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + 1)
    cell = ws.cell(row=row, column=col)
    cell.fill = KPI_FILL
    cell.border = BORDER
    if fmt == "currency":
        cell.value = value
        cell.number_format = '"\u20b9"#,##0'
    else:
        cell.value = value
    cell.font = KPI_VALUE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
    label_cell = ws.cell(row=row + 2, column=col, value=label)
    label_cell.font = KPI_LABEL_FONT
    label_cell.alignment = Alignment(horizontal="center")
    for rr in (row, row + 1, row + 2):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).fill = KPI_FILL


def build_report(period: str):
    sales = pd.read_csv("data/processed/sales_cleaned.csv", parse_dates=["order_date"])
    max_date = sales["order_date"].max()

    if period == "weekly":
        cutoff = max_date - pd.Timedelta(days=7)
        label = f"Week of {cutoff.date()} to {max_date.date()}"
        fname = f"reports/Weekly_Sales_Report_{max_date.date()}.xlsx"
    else:
        cutoff = max_date - pd.Timedelta(days=30)
        label = f"{max_date.strftime('%B %Y')}"
        fname = f"reports/Monthly_Sales_Report_{max_date.strftime('%Y_%m')}.xlsx"

    period_df = sales[sales["order_date"] > cutoff]
    prior_window = sales[(sales["order_date"] <= cutoff) &
                          (sales["order_date"] > cutoff - (max_date - cutoff))]

    total_revenue = period_df["revenue"].sum()
    prior_revenue = prior_window["revenue"].sum()
    pct_change = ((total_revenue - prior_revenue) / prior_revenue * 100) if prior_revenue else 0
    num_orders = len(period_df)
    avg_order_value = period_df["revenue"].mean() if num_orders else 0

    region_summary = (period_df.groupby("region")["revenue"].sum()
                       .reset_index().sort_values("revenue", ascending=False)
                       .rename(columns={"revenue": "total_revenue"}))
    category_summary = (period_df.groupby("product_category")
                         .agg(units_sold=("quantity", "sum"), total_revenue=("revenue", "sum"))
                         .reset_index().sort_values("total_revenue", ascending=False))
    top_products = (period_df.groupby("product_name")
                     .agg(units_sold=("quantity", "sum"), total_revenue=("revenue", "sum"))
                     .reset_index().sort_values("total_revenue", ascending=False).head(5))

    daily_trend = (sales[sales["order_date"] > max_date - pd.Timedelta(days=30)]
                    .groupby(sales["order_date"].dt.date)["revenue"].sum().reset_index())
    daily_trend.columns = ["date", "revenue"]

    # ---------------- BUILD WORKBOOK -----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = f"SalesPulse {period.capitalize()} Sales Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:H2")
    ws["A2"] = f"{label}   |   Generated automatically on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = SUBTITLE_FONT

    add_kpi_card(ws, 1, 4, "Total Revenue", round(total_revenue), fmt="currency")
    add_kpi_card(ws, 3, 4, "Orders", num_orders)
    add_kpi_card(ws, 5, 4, "Avg Order Value", round(avg_order_value), fmt="currency")
    pct_cell_row = 4
    ws.merge_cells(start_row=pct_cell_row, start_column=7, end_row=pct_cell_row + 1, end_column=8)
    chg_cell = ws.cell(row=pct_cell_row, column=7, value=f"{pct_change:+.1f}%")
    chg_cell.font = Font(name="Calibri", bold=True, size=20,
                          color="2E7D32" if pct_change >= 0 else "C62828")
    chg_cell.alignment = Alignment(horizontal="center", vertical="center")
    chg_cell.fill = KPI_FILL
    ws.merge_cells(start_row=pct_cell_row + 2, start_column=7, end_row=pct_cell_row + 2, end_column=8)
    lbl = ws.cell(row=pct_cell_row + 2, column=7, value="vs. Prior Period")
    lbl.font = KPI_LABEL_FONT
    lbl.alignment = Alignment(horizontal="center")
    for rr in (4, 5, 6):
        for cc in (7, 8):
            ws.cell(row=rr, column=cc).fill = KPI_FILL

    row = 8
    ws.cell(row=row, column=1, value="Revenue by Region").font = LABEL_FONT
    row += 1
    region_end = write_table(ws, region_summary, row, currency_cols=["total_revenue"])

    row = region_end + 2
    ws.cell(row=row, column=1, value="Revenue by Category").font = LABEL_FONT
    row += 1
    cat_end = write_table(ws, category_summary, row, currency_cols=["total_revenue"])

    row = cat_end + 2
    ws.cell(row=row, column=1, value="Top 5 Products").font = LABEL_FONT
    row += 1
    top_end = write_table(ws, top_products, row, currency_cols=["total_revenue"])

    # Daily trend sheet + native Excel chart
    ws2 = wb.create_sheet("Trend (Last 30 Days)")
    ws2.sheet_view.showGridLines = False
    ws2.cell(row=1, column=1, value="Date").font = HEADER_FONT
    ws2.cell(row=1, column=2, value="Revenue").font = HEADER_FONT
    ws2.cell(row=1, column=1).fill = HEADER_FILL
    ws2.cell(row=1, column=2).fill = HEADER_FILL
    for i, r in daily_trend.iterrows():
        ws2.cell(row=i + 2, column=1, value=str(r["date"]))
        ws2.cell(row=i + 2, column=2, value=round(r["revenue"]))

    chart = LineChart()
    chart.title = "Daily Revenue Trend"
    chart.style = 2
    chart.y_axis.title = "Revenue (\u20b9)"
    chart.x_axis.title = "Date"
    data_ref = Reference(ws2, min_col=2, min_row=1, max_row=1 + len(daily_trend))
    cats_ref = Reference(ws2, min_col=1, min_row=2, max_row=1 + len(daily_trend))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width, chart.height = 24, 12
    ws2.add_chart(chart, "D2")

    for col, width in zip("ABCDEFGH", [20, 14, 14, 14, 14, 14, 14, 14]):
        ws.column_dimensions[col].width = width
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.page_margins.left = 0.4
    ws2.page_margins.right = 0.4

    wb.save(fname)
    print(f"Report saved: {fname}")
    print(f"  Period: {label}")
    print(f"  Total revenue: \u20b9{total_revenue:,.0f}  ({pct_change:+.1f}% vs prior period)")
    print(f"  Orders: {num_orders:,}  |  Avg order value: \u20b9{avg_order_value:,.0f}")
    return fname


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--period", choices=["weekly", "monthly"], default="monthly")
    args = parser.parse_args()
    build_report(args.period)
